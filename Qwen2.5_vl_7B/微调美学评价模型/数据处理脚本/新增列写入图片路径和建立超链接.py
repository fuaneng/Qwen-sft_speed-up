# import os
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.styles import Font
# from openpyxl.utils import get_column_letter

# def update_image_paths_with_hyperlinks(excel_path, sheet_name, image_folder_path):
#     """
#     根据Excel中的图片名，在指定文件夹中查找对应的图片路径，
#     并以绝对路径超链接格式写入Excel，以避免复制后链接失效的问题。

#     Args:
#         excel_path (str): Excel文件的完整路径。
#         sheet_name (str): 包含图片名的工作表名称。
#         image_folder_path (str): 包含所有图片的文件夹路径。
#     """
#     try:
#         # 1. 准备数据：使用 Pandas 获取数据和路径
#         df = pd.read_excel(excel_path, sheet_name=sheet_name)
#         print(f"成功加载Excel文件：'{excel_path}'")

#         if '图片名' not in df.columns:
#             print("错误：Excel文件中未找到名为 '图片名' 的列。")
#             return

#         # 确保图片文件夹路径是绝对路径
#         image_folder_path = os.path.abspath(image_folder_path)
#         if not os.path.isdir(image_folder_path):
#             print(f"错误：图片文件夹路径不存在或不是一个目录：'{image_folder_path}'")
#             return

#         # 创建一个字典，键为不含后缀的图片名，值为完整绝对路径
#         image_files = {}
#         for file_name in os.listdir(image_folder_path):
#             # 使用 os.path.join() 和 os.path.abspath() 确保路径是绝对路径
#             full_path = os.path.abspath(os.path.join(image_folder_path, file_name))
#             if os.path.isfile(full_path):
#                 file_name_without_ext, _ = os.path.splitext(file_name)
#                 image_files[file_name_without_ext] = full_path

#         df['图片路径'] = ''
#         for index, row in df.iterrows():
#             excel_image_name = str(row['图片名'])
#             excel_name_without_ext, _ = os.path.splitext(excel_image_name)

#             if excel_name_without_ext in image_files:
#                 df.at[index, '图片路径'] = image_files[excel_name_without_ext]
#             else:
#                 print(f"警告：未在图片文件夹中找到与 '{excel_image_name}' 匹配的图片文件。")

#         # 2. 写入超链接到 Excel
#         workbook = load_workbook(excel_path)
#         if sheet_name in workbook.sheetnames:
#             worksheet = workbook[sheet_name]
#         else:
#             print(f"错误：Excel文件中未找到名为 '{sheet_name}' 的工作表。")
#             return

#         column_headers = list(df.columns)
#         for col_index, header in enumerate(column_headers, 1):
#             worksheet.cell(row=1, column=col_index, value=header)

#         for r_idx, row_data in enumerate(df.itertuples(index=False), 2):
#             for c_idx, cell_value in enumerate(row_data, 1):
#                 if column_headers[c_idx - 1] == '图片路径':
#                     if cell_value:
#                         # 转换路径为URL格式，确保是绝对路径
#                         hyperlink_path = 'file:///' + os.path.normpath(cell_value).replace('\\', '/')

#                         cell = worksheet.cell(row=r_idx, column=c_idx, value=cell_value)
#                         cell.hyperlink = hyperlink_path
#                         cell.font = Font(color="0000FF", underline="single")
#                 else:
#                     worksheet.cell(row=r_idx, column=c_idx, value=cell_value)

#         # 调整列宽
#         for col in worksheet.columns:
#             max_length = 0
#             column = [cell for cell in col]
#             for cell in column:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass
#             adjusted_width = (max_length + 2)
#             worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

#         # 3. 保存更新后的Excel
#         workbook.save(excel_path)
#         print(f"任务完成！更新后的数据（含绝对路径超链接）已保存到 '{excel_path}'")

#     except FileNotFoundError:
#         print(f"错误：文件或目录未找到。请检查路径是否正确。\nExcel路径: {excel_path}\n图片文件夹路径: {image_folder_path}")
#     except Exception as e:
#         print(f"发生了一个错误：{e}")

# # --- 配置你的文件路径和参数 ---
# excel_file_path = r'D:\work\tain\美学评价\data\set\训练数据_美学评测结果.xlsx'
# sheet_name_to_update = 'Sheet1'
# images_folder = r'D:\work\tain\美学评价\data\set\训练数据\512'

# # --- 运行脚本 ---
# update_image_paths_with_hyperlinks(excel_file_path, sheet_name_to_update, images_folder)


# ##---在所有文件夹下查找图片---##
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def update_image_paths_with_hyperlinks(excel_path, sheet_name, image_folder_path):
    """
    根据Excel中的图片名，在指定文件夹及其所有子文件夹中查找对应的图片路径，
    并以绝对路径超链接格式写入Excel。

    Args:
        excel_path (str): Excel文件的完整路径。
        sheet_name (str): 包含图片名的工作表名称。
        image_folder_path (str): 包含所有图片的根文件夹路径。
    """
    try:
        # 1. 准备数据：使用 Pandas 获取数据和路径
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        print(f"成功加载Excel文件：'{excel_path}'")

        if '图片名' not in df.columns:
            print("错误：Excel文件中未找到名为 '图片名' 的列。")
            return

        image_folder_path = os.path.abspath(image_folder_path)
        if not os.path.isdir(image_folder_path):
            print(f"错误：图片文件夹路径不存在或不是一个目录：'{image_folder_path}'")
            return

        # 创建一个字典，键为不含后缀的图片名，值为完整绝对路径
        image_files = {}
        
        # --- 关键修改：使用 os.walk() 递归遍历子文件夹 ---
        print(f"正在搜索 '{image_folder_path}' 及其所有子文件夹中的图片...")
        for root, _, files in os.walk(image_folder_path):
            for file_name in files:
                # 使用 os.path.join() 构建完整路径
                full_path = os.path.join(root, file_name)
                
                # 排除非图片文件（可选，但推荐）
                if file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                    file_name_without_ext, _ = os.path.splitext(file_name)
                    # 将不含后缀的图片名作为键，完整绝对路径作为值
                    image_files[file_name_without_ext] = full_path
        
        print(f"共找到 {len(image_files)} 个图片文件。")
        # --- 关键修改结束 ---
        
        df['图片路径'] = ''
        for index, row in df.iterrows():
            # 确保 Excel 中的图片名不包含后缀，与我们的字典键匹配
            excel_image_name = str(row['图片名'])
            excel_name_without_ext, _ = os.path.splitext(excel_image_name)

            if excel_name_without_ext in image_files:
                df.at[index, '图片路径'] = image_files[excel_name_without_ext]
            else:
                print(f"警告：未在任何子文件夹中找到与 '{excel_image_name}' 匹配的图片文件。")

        # 2. 写入超链接到 Excel
        workbook = load_workbook(excel_path)
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            print(f"错误：Excel文件中未找到名为 '{sheet_name}' 的工作表。")
            return

        column_headers = list(df.columns)
        # 清空并重写表头，确保一致性
        for col_index, header in enumerate(column_headers, 1):
            worksheet.cell(row=1, column=col_index, value=header)

        for r_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for c_idx, cell_value in enumerate(row_data, 1):
                if column_headers[c_idx - 1] == '图片路径':
                    if cell_value:
                        # 转换路径为URL格式，确保是绝对路径
                        hyperlink_path = 'file:///' + os.path.normpath(cell_value).replace('\\', '/')

                        cell = worksheet.cell(row=r_idx, column=c_idx, value=cell_value)
                        cell.hyperlink = hyperlink_path
                        cell.font = Font(color="0000FF", underline="single")
                    else:
                        # 如果未找到图片，写入一个空字符串或自定义文本
                        worksheet.cell(row=r_idx, column=c_idx, value="未找到图片")
                else:
                    worksheet.cell(row=r_idx, column=c_idx, value=cell_value)

        # 调整列宽
        for col in worksheet.columns:
            max_length = 0
            column = [cell for cell in col]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
        # 3. 保存更新后的Excel
        workbook.save(excel_path)
        print(f"任务完成！更新后的数据（含绝对路径超链接）已保存到 '{excel_path}'")

    except FileNotFoundError:
        print(f"错误：文件或目录未找到。请检查路径是否正确。\nExcel路径: {excel_path}\n图片文件夹路径: {image_folder_path}")
    except Exception as e:
        print(f"发生了一个错误：{e}")

# --- 配置你的文件路径和参数 ---
excel_file_path = r'D:\work\tain\美学评价\data\set\训练数据_美学评测结果.xlsx'
sheet_name_to_update = 'Sheet1'

# 新的图片根文件夹，现在脚本会在此文件夹下的所有子文件夹中搜索图片
images_folder = r'D:\work\tain\美学评价\data\set\低分段'

# --- 运行脚本 ---
update_image_paths_with_hyperlinks(excel_file_path, sheet_name_to_update, images_folder)