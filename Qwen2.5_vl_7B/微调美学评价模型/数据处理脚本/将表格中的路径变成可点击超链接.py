import openpyxl
import os

def create_hyperlinks_for_multiple_columns(file_path, column_names):
    """
    将指定Excel文件中所有工作表中的多个指定列文本内容转换为超链接。

    Args:
        file_path (str): Excel文件的完整路径。
        column_names (list): 一个包含需要转换的列名的列表。
    """
    try:
        # 1. 打开Excel文件
        workbook = openpyxl.load_workbook(file_path)
        
        # 2. 遍历所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            print(f"正在处理 '{sheet_name}' 表...")
            
            # 3. 遍历需要修改的每个列
            for column_name in column_names:
                # 4. 查找目标列的索引
                # 假设第一行是列标题
                header_row = sheet[1]
                column_index = -1
                for cell in header_row:
                    if cell.value == column_name:
                        column_index = cell.column
                        break
                
                if column_index == -1:
                    print(f"  警告: 在工作表 '{sheet_name}' 中找不到列 '{column_name}'，已跳过。")
                    continue

                print(f"  正在处理 '{column_name}' 列...")
                
                # 5. 遍历目标列，创建超链接
                for row in range(2, sheet.max_row + 1):  # 从第二行开始处理数据
                    cell = sheet.cell(row=row, column=column_index)
                    
                    if cell.value and isinstance(cell.value, str):
                        file_path_to_link = cell.value.strip()
                        
                        if os.path.exists(file_path_to_link):
                            processed_path = file_path_to_link.replace("\\", "/")
                            hyperlink_url = f'file:///{processed_path}'
                            
                            cell.hyperlink = hyperlink_url
                            cell.value = file_path_to_link
                        else:
                            print(f"    警告: 在 '{sheet_name}' 表的第 {row} 行，路径不存在: {file_path_to_link}")
        
        # 6. 保存修改后的文件
        new_file_path = os.path.splitext(file_path)[0] + "_带多列超链接.xlsx"
        workbook.save(new_file_path)
        print(f"\n文件处理完成！已保存为: {new_file_path}")

    except FileNotFoundError:
        print(f"错误: 找不到文件: {file_path}")
    except Exception as e:
        print(f"处理过程中发生错误: {e}")

# --- 配置参数 ---
# 请根据你的文件路径和列名修改以下变量
excel_file_path = r'D:\work\tain\美学评价\data\xlsx\汇总.xlsx'
# 将所有需要转换的列名放入一个列表中
columns_to_process = ['图片路径', '原图_图片路径']

# --- 调用函数 ---
create_hyperlinks_for_multiple_columns(excel_file_path, columns_to_process)